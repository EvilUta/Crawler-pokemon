import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

url = 'https://pokemondb.net/pokedex/all'

r = requests.get(url)
r.encoding = 'utf-8'
soup = BeautifulSoup(r.text, 'html.parser')

table = soup.find('table', {'id': 'pokedex'})
rows = table.tbody.find_all('tr')

dados = []
for row in rows:
    cols = row.find_all('td')
    numero = cols[0].text.strip()
    nome = cols[1].text.strip()
    tipos = ', '.join([t.text for t in cols[2].find_all('a')])
    total = cols[3].text.strip()
    hp = cols[4].text.strip()
    ataque = cols[5].text.strip()
    defesa = cols[6].text.strip()
    atk_esp = cols[7].text.strip()
    def_esp = cols[8].text.strip()
    velocidade = cols[9].text.strip()

    dados.append({
        'Número': numero,
        'Nome': nome,
        'Tipo(s)': tipos,
        'Total': total,
        'HP': hp,
        'Ataque': ataque,
        'Defesa': defesa,
        'Atk. Esp.': atk_esp,
        'Def. Esp.': def_esp,
        'Velocidade': velocidade
    })

# Cria DataFrame
df = pd.DataFrame(dados)

arquivo = 'pokemondb_pokedex.xlsx'
df.to_excel(arquivo, index=False)

# Formatação da planilha
wb = load_workbook(arquivo)
ws = wb.active

# Ajusta largura colunas
larguras = [10, 25, 20, 10, 8, 8, 8, 10, 10, 12]
colunas = ['A','B','C','D','E','F','G','H','I','J']
for col, largura in zip(colunas, larguras):
    ws.column_dimensions[col].width = largura

# Estilo cabeçalho
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill("solid", fgColor="4F81BD")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Alinha colunas numéricas no centro
for row in ws.iter_rows(min_row=2, min_col=4, max_col=10):
    for cell in row:
        cell.alignment = Alignment(horizontal='center')

wb.save(arquivo)

print(f'Planilha "{arquivo}" criada com {len(df)} Pokémon com layout formatado.')
